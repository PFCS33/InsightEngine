import openpyxl
import os


class DataSource:
    def __init__(self, source_name, source_path, req_id):
        # filename without suffix
        self.source_name = source_name
        # filename suffix
        self.name_suffix = source_name.split('.')[1]
        # /data/filename
        self.source_path = source_path
        self.request_uuid = req_id
        # /results/req_id/filename.json
        self.result_path = self.get_result_path()
        # self.get_header_range()

    def get_result_path(self):
        # current_date = datetime.date.today().strftime("%m%d")
        # path = 'data/' + self.source_name + '/'
        # path = path + 'result_' + current_date + '_S' + str(state) + '.json'
        folder = 'results/' + self.request_uuid
        if not os.path.exists(folder):
            os.makedirs(folder)
        path = folder + '/' + self.source_name + '.json'
        return path

    def get_state_data_path(self):
        folder = 'data'
        if not os.path.exists(folder):
            os.makedirs(folder)
        path = folder + "/" + self.source_name + '.xlsx'
        return path

    # def get_header_range(self):
    #     table = openpyxl.load_workbook(self.source_path).active
    #     header = []  # header range
    #     for i in range(table.max_row):
    #         value = table.cell(row=i+1, column=1).value
    #         if value is None or value.isspace():
    #             header.append(i)
    #         else:
    #             break

    #     index = []  # index range
    #     for i in range(table.max_column):
    #         value = table.cell(row=1, column=i+1).value
    #         if value is None or value.isspace():
    #             index.append(i)
    #         else:
    #             break

    #     self.header_row = header
    #     self.header_col = index
